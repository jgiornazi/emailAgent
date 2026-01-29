"""
Excel Storage Module

This module handles all Excel operations for storing job application data:
- Creating and loading Excel files
- Adding new company rows
- Updating existing company rows
- Status hierarchy enforcement
- Conflict detection and flagging
- Auto-backup functionality
- Export to CSV/JSON

File Structure:
    Column A: Company Name
    Column B: Position
    Column C: Status (Applied, Interviewing, Rejected, Offer)
    Column D: Confidence (high, medium, low)
    Column E: Date First Seen
    Column F: Date Last Updated
    Column G: Email IDs (comma-separated)
    Column H: Notes
"""

import csv
import json
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle

from .extractor import ExtractionResult
from .classifier import (
    can_update_status,
    create_conflict_note,
    get_status_level,
    STATUS_HIERARCHY,
)


# =============================================================================
# Constants
# =============================================================================

SHEET_NAME = "Applications"

# Column configuration (1-based indices)
COLUMNS = {
    'company': 1,      # A
    'position': 2,     # B
    'status': 3,       # C
    'confidence': 4,   # D
    'date_first': 5,   # E
    'date_last': 6,    # F
    'email_ids': 7,    # G
    'notes': 8,        # H
}

HEADERS = [
    "Company Name",
    "Position",
    "Status",
    "Confidence",
    "Date First Seen",
    "Date Last Updated",
    "Email IDs",
    "Notes",
]

# Column widths
COLUMN_WIDTHS = {
    1: 25,   # Company Name
    2: 30,   # Position
    3: 15,   # Status
    4: 12,   # Confidence
    5: 15,   # Date First Seen
    6: 18,   # Date Last Updated
    7: 40,   # Email IDs
    8: 50,   # Notes
}

# Styling
HEADER_FILL = PatternFill(start_color="D5E8F0", end_color="D5E8F0", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Status colors for conditional formatting
STATUS_COLORS = {
    'Applied': PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),      # Light blue
    'Interviewing': PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),  # Light yellow
    'Rejected': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),      # Light red
    'Offer': PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),         # Light green
}

CONFLICT_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Red highlight


# =============================================================================
# Data Classes
# =============================================================================

@dataclass
class JobApplication:
    """Represents a job application row in Excel."""

    company: str
    position: str = "Not specified"
    status: str = "Applied"
    confidence: str = "medium"
    date_first: Optional[datetime] = None
    date_last: Optional[datetime] = None
    email_ids: List[str] = None
    notes: str = ""
    row_index: Optional[int] = None  # Excel row number (1-based)

    def __post_init__(self):
        if self.email_ids is None:
            self.email_ids = []
        if self.date_first is None:
            self.date_first = datetime.now()
        if self.date_last is None:
            self.date_last = datetime.now()

    @property
    def has_conflict(self) -> bool:
        """Check if this application has a conflict flag."""
        return "Conflict:" in self.notes

    def add_email_id(self, email_id: str) -> None:
        """Add an email ID if not already present."""
        if email_id and email_id not in self.email_ids:
            self.email_ids.append(email_id)

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            'company': self.company,
            'position': self.position,
            'status': self.status,
            'confidence': self.confidence,
            'date_first': self.date_first.isoformat() if self.date_first else None,
            'date_last': self.date_last.isoformat() if self.date_last else None,
            'email_ids': self.email_ids,
            'notes': self.notes,
        }


@dataclass
class ExcelUpdateResult:
    """Result of an Excel update operation."""

    success: bool
    is_new_row: bool = False
    is_update: bool = False
    is_conflict: bool = False
    row_index: int = 0
    company: str = ""
    old_status: Optional[str] = None
    new_status: Optional[str] = None
    message: str = ""


# =============================================================================
# Excel Storage Class
# =============================================================================

class ExcelStorage:
    """
    Manages Excel file operations for job applications.

    Usage:
        storage = ExcelStorage("~/job_applications.xlsx")
        storage.initialize()

        # Add or update application
        result = storage.add_or_update(extraction_result)

        # Save changes
        storage.save()
    """

    def __init__(
        self,
        file_path: str = "~/job_applications.xlsx",
        backup_dir: str = "~/.emailagent/backups/",
        auto_backup: bool = True,
        backup_retention_days: int = 7,
    ):
        """
        Initialize Excel storage.

        Args:
            file_path: Path to Excel file
            backup_dir: Directory for backups
            auto_backup: Whether to create backups automatically
            backup_retention_days: Days to keep old backups
        """
        self.file_path = Path(file_path).expanduser()
        self.backup_dir = Path(backup_dir).expanduser()
        self.auto_backup = auto_backup
        self.backup_retention_days = backup_retention_days

        self.workbook: Optional[Workbook] = None
        self.worksheet = None
        self._company_cache: Dict[str, int] = {}  # company_name -> row_index
        self._modified = False
        self._unsaved_count = 0

    def initialize(self) -> None:
        """
        Initialize the Excel file and load or create workbook.

        Creates the file with headers if it doesn't exist.
        """
        self.backup_dir.mkdir(parents=True, exist_ok=True)

        if self.file_path.exists():
            # Create backup before loading
            if self.auto_backup:
                self._create_backup()

            # Load existing workbook
            self.workbook = load_workbook(self.file_path)

            # Get or create the Applications sheet
            if SHEET_NAME in self.workbook.sheetnames:
                self.worksheet = self.workbook[SHEET_NAME]
            else:
                self.worksheet = self.workbook.create_sheet(SHEET_NAME)
                self._setup_headers()

            # Build company cache
            self._build_company_cache()
        else:
            # Create new workbook
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = SHEET_NAME
            self._setup_headers()
            self._company_cache = {}

        self._modified = False
        self._unsaved_count = 0

    def _setup_headers(self) -> None:
        """Set up header row with formatting."""
        # Write headers
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = self.worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

        # Set column widths
        for col_idx, width in COLUMN_WIDTHS.items():
            col_letter = get_column_letter(col_idx)
            self.worksheet.column_dimensions[col_letter].width = width

        # Freeze header row
        self.worksheet.freeze_panes = "A2"

        # Add thin border to header
        thin_border = Border(
            bottom=Side(style='thin', color='808080')
        )
        for col_idx in range(1, len(HEADERS) + 1):
            self.worksheet.cell(row=1, column=col_idx).border = thin_border

    def _build_company_cache(self) -> None:
        """Build cache of company names to row indices."""
        self._company_cache = {}

        for row_idx in range(2, self.worksheet.max_row + 1):
            company_cell = self.worksheet.cell(row=row_idx, column=COLUMNS['company'])
            if company_cell.value:
                company_lower = company_cell.value.lower().strip()
                self._company_cache[company_lower] = row_idx

    def _create_backup(self) -> Optional[Path]:
        """Create a timestamped backup of the Excel file."""
        if not self.file_path.exists():
            return None

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"job_applications_backup_{timestamp}.xlsx"
        backup_path = self.backup_dir / backup_name

        shutil.copy2(self.file_path, backup_path)

        # Clean old backups
        self._cleanup_old_backups()

        return backup_path

    def _cleanup_old_backups(self) -> int:
        """Remove backups older than retention period."""
        if not self.backup_dir.exists():
            return 0

        cutoff = datetime.now().timestamp() - (self.backup_retention_days * 86400)
        removed_count = 0

        for backup_file in self.backup_dir.glob("job_applications_backup_*.xlsx"):
            if backup_file.stat().st_mtime < cutoff:
                backup_file.unlink()
                removed_count += 1

        return removed_count

    def find_company(self, company_name: str) -> Optional[int]:
        """
        Find the row index for a company (case-insensitive).

        Args:
            company_name: Company name to search for

        Returns:
            Row index (1-based) or None if not found
        """
        company_lower = company_name.lower().strip()
        return self._company_cache.get(company_lower)

    def get_application(self, row_index: int) -> Optional[JobApplication]:
        """
        Get application data from a row.

        Args:
            row_index: Excel row number (1-based)

        Returns:
            JobApplication or None if row is empty
        """
        if row_index < 2 or row_index > self.worksheet.max_row:
            return None

        company = self.worksheet.cell(row=row_index, column=COLUMNS['company']).value
        if not company:
            return None

        # Parse dates
        date_first_val = self.worksheet.cell(row=row_index, column=COLUMNS['date_first']).value
        date_last_val = self.worksheet.cell(row=row_index, column=COLUMNS['date_last']).value

        date_first = None
        date_last = None

        if date_first_val:
            if isinstance(date_first_val, datetime):
                date_first = date_first_val
            elif isinstance(date_first_val, str):
                try:
                    date_first = datetime.fromisoformat(date_first_val)
                except ValueError:
                    pass

        if date_last_val:
            if isinstance(date_last_val, datetime):
                date_last = date_last_val
            elif isinstance(date_last_val, str):
                try:
                    date_last = datetime.fromisoformat(date_last_val)
                except ValueError:
                    pass

        # Parse email IDs
        email_ids_str = self.worksheet.cell(row=row_index, column=COLUMNS['email_ids']).value
        email_ids = []
        if email_ids_str:
            email_ids = [eid.strip() for eid in email_ids_str.split(',') if eid.strip()]

        return JobApplication(
            company=company,
            position=self.worksheet.cell(row=row_index, column=COLUMNS['position']).value or "Not specified",
            status=self.worksheet.cell(row=row_index, column=COLUMNS['status']).value or "Applied",
            confidence=self.worksheet.cell(row=row_index, column=COLUMNS['confidence']).value or "medium",
            date_first=date_first,
            date_last=date_last,
            email_ids=email_ids,
            notes=self.worksheet.cell(row=row_index, column=COLUMNS['notes']).value or "",
            row_index=row_index,
        )

    def add_new_row(self, extraction: ExtractionResult) -> ExcelUpdateResult:
        """
        Add a new company row to Excel.

        Args:
            extraction: Extraction result with company, position, status, etc.

        Returns:
            ExcelUpdateResult with operation details
        """
        # Get next row
        next_row = self.worksheet.max_row + 1

        # Format date
        date_str = extraction.email_date.strftime('%Y-%m-%d') if extraction.email_date else datetime.now().strftime('%Y-%m-%d')

        # Write data
        self.worksheet.cell(row=next_row, column=COLUMNS['company'], value=extraction.company)
        self.worksheet.cell(row=next_row, column=COLUMNS['position'], value=extraction.position)
        self.worksheet.cell(row=next_row, column=COLUMNS['status'], value=extraction.status)
        self.worksheet.cell(row=next_row, column=COLUMNS['confidence'], value=extraction.confidence)
        self.worksheet.cell(row=next_row, column=COLUMNS['date_first'], value=date_str)
        self.worksheet.cell(row=next_row, column=COLUMNS['date_last'], value=date_str)
        self.worksheet.cell(row=next_row, column=COLUMNS['email_ids'], value=extraction.email_id)
        self.worksheet.cell(row=next_row, column=COLUMNS['notes'], value="")

        # Apply status color
        self._apply_status_formatting(next_row, extraction.status)

        # Mark as low confidence if needed
        if extraction.confidence == 'low':
            notes_cell = self.worksheet.cell(row=next_row, column=COLUMNS['notes'])
            notes_cell.value = "NEEDS REVIEW"

        # Update cache
        self._company_cache[extraction.company.lower().strip()] = next_row

        self._modified = True
        self._unsaved_count += 1

        return ExcelUpdateResult(
            success=True,
            is_new_row=True,
            row_index=next_row,
            company=extraction.company,
            new_status=extraction.status,
            message=f"Added new row for {extraction.company}",
        )

    def update_existing_row(
        self,
        row_index: int,
        extraction: ExtractionResult,
    ) -> ExcelUpdateResult:
        """
        Update an existing company row.

        Handles status hierarchy and conflict detection.

        Args:
            row_index: Excel row number to update
            extraction: New extraction result

        Returns:
            ExcelUpdateResult with operation details
        """
        # Get current data
        current = self.get_application(row_index)
        if not current:
            return ExcelUpdateResult(
                success=False,
                message=f"Row {row_index} not found",
            )

        # Check status hierarchy
        update_result = can_update_status(current.status, extraction.status)

        # Format date
        date_str = extraction.email_date.strftime('%Y-%m-%d') if extraction.email_date else datetime.now().strftime('%Y-%m-%d')

        if update_result.allowed:
            # Update status
            self.worksheet.cell(row=row_index, column=COLUMNS['status'], value=extraction.status)

            # Update position if new one is more specific
            if extraction.position != "Not specified":
                self.worksheet.cell(row=row_index, column=COLUMNS['position'], value=extraction.position)

            # Update confidence if higher
            confidence_order = {'low': 0, 'medium': 1, 'high': 2}
            if confidence_order.get(extraction.confidence, 0) > confidence_order.get(current.confidence, 0):
                self.worksheet.cell(row=row_index, column=COLUMNS['confidence'], value=extraction.confidence)

            # Update last date
            self.worksheet.cell(row=row_index, column=COLUMNS['date_last'], value=date_str)

            # Append email ID
            self._append_email_id(row_index, extraction.email_id)

            # Apply status formatting
            self._apply_status_formatting(row_index, extraction.status)

            self._modified = True
            self._unsaved_count += 1

            return ExcelUpdateResult(
                success=True,
                is_update=True,
                row_index=row_index,
                company=current.company,
                old_status=current.status,
                new_status=extraction.status,
                message=f"Updated {current.company}: {current.status} -> {extraction.status}",
            )
        else:
            # Conflict - do not update status, but add note and track email
            self._handle_conflict(
                row_index,
                current.status,
                extraction.status,
                date_str,
                extraction.email_id,
            )

            self._modified = True
            self._unsaved_count += 1

            return ExcelUpdateResult(
                success=True,
                is_conflict=True,
                row_index=row_index,
                company=current.company,
                old_status=current.status,
                new_status=extraction.status,
                message=f"CONFLICT at {current.company}: kept {current.status}, received {extraction.status}",
            )

    def _handle_conflict(
        self,
        row_index: int,
        current_status: str,
        new_status: str,
        date_str: str,
        email_id: str,
    ) -> None:
        """Handle a status conflict by flagging in Excel."""
        # Create conflict note
        conflict_note = f"Conflict: received {new_status} after {current_status} on {date_str}"

        # Get current notes
        notes_cell = self.worksheet.cell(row=row_index, column=COLUMNS['notes'])
        current_notes = notes_cell.value or ""

        # Append conflict note
        if current_notes:
            updated_notes = f"{current_notes}; {conflict_note}"
        else:
            updated_notes = conflict_note

        notes_cell.value = updated_notes

        # Apply conflict highlighting
        notes_cell.fill = CONFLICT_FILL

        # Still append email ID (track conflicting email)
        self._append_email_id(row_index, email_id)

        # Update last date (when conflict occurred)
        self.worksheet.cell(row=row_index, column=COLUMNS['date_last'], value=date_str)

    def _append_email_id(self, row_index: int, email_id: str) -> None:
        """Append an email ID to the existing list."""
        if not email_id:
            return

        ids_cell = self.worksheet.cell(row=row_index, column=COLUMNS['email_ids'])
        current_ids = ids_cell.value or ""

        # Check if already present
        if email_id in current_ids:
            return

        # Append
        if current_ids:
            ids_cell.value = f"{current_ids}, {email_id}"
        else:
            ids_cell.value = email_id

    def _apply_status_formatting(self, row_index: int, status: str) -> None:
        """Apply color formatting based on status."""
        status_fill = STATUS_COLORS.get(status)
        if status_fill:
            status_cell = self.worksheet.cell(row=row_index, column=COLUMNS['status'])
            status_cell.fill = status_fill

    def add_or_update(self, extraction: ExtractionResult) -> ExcelUpdateResult:
        """
        Add new row or update existing row for a company.

        This is the main entry point for processing extraction results.

        Args:
            extraction: Extraction result from email processing

        Returns:
            ExcelUpdateResult with operation details
        """
        # Find existing row
        existing_row = self.find_company(extraction.company)

        if existing_row:
            return self.update_existing_row(existing_row, extraction)
        else:
            return self.add_new_row(extraction)

    def save(self, force: bool = False) -> bool:
        """
        Save the workbook to file.

        Args:
            force: Save even if no modifications

        Returns:
            True if saved successfully
        """
        if not self._modified and not force:
            return True

        try:
            # Ensure directory exists
            self.file_path.parent.mkdir(parents=True, exist_ok=True)

            # Save workbook
            self.workbook.save(self.file_path)

            self._modified = False
            self._unsaved_count = 0

            return True
        except Exception as e:
            raise IOError(f"Failed to save Excel file: {e}")

    def save_if_needed(self, threshold: int = 100) -> bool:
        """
        Save if unsaved changes exceed threshold.

        Args:
            threshold: Number of unsaved changes before auto-save

        Returns:
            True if saved
        """
        if self._unsaved_count >= threshold:
            return self.save()
        return False

    def get_all_applications(self) -> List[JobApplication]:
        """
        Get all applications from Excel.

        Returns:
            List of JobApplication objects
        """
        applications = []

        for row_idx in range(2, self.worksheet.max_row + 1):
            app = self.get_application(row_idx)
            if app:
                applications.append(app)

        return applications

    def get_applications_by_status(self, status: str) -> List[JobApplication]:
        """
        Get applications filtered by status.

        Args:
            status: Status to filter by

        Returns:
            List of matching applications
        """
        return [app for app in self.get_all_applications() if app.status == status]

    def get_conflicts(self) -> List[JobApplication]:
        """
        Get all applications with conflicts.

        Returns:
            List of applications with conflict flags
        """
        return [app for app in self.get_all_applications() if app.has_conflict]

    def get_statistics(self) -> Dict[str, Any]:
        """
        Get summary statistics.

        Returns:
            Dictionary with counts and statistics
        """
        all_apps = self.get_all_applications()

        status_counts = {status: 0 for status in STATUS_HIERARCHY}
        confidence_counts = {'high': 0, 'medium': 0, 'low': 0}
        conflict_count = 0

        for app in all_apps:
            if app.status in status_counts:
                status_counts[app.status] += 1
            if app.confidence in confidence_counts:
                confidence_counts[app.confidence] += 1
            if app.has_conflict:
                conflict_count += 1

        return {
            'total_companies': len(all_apps),
            'status_counts': status_counts,
            'confidence_counts': confidence_counts,
            'conflict_count': conflict_count,
        }

    def export_to_csv(self, output_path: str) -> str:
        """
        Export data to CSV file.

        Args:
            output_path: Path for CSV file

        Returns:
            Path to created file
        """
        output = Path(output_path).expanduser()

        with open(output, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Write headers
            writer.writerow(HEADERS)

            # Write data
            for app in self.get_all_applications():
                writer.writerow([
                    app.company,
                    app.position,
                    app.status,
                    app.confidence,
                    app.date_first.strftime('%Y-%m-%d') if app.date_first else '',
                    app.date_last.strftime('%Y-%m-%d') if app.date_last else '',
                    ', '.join(app.email_ids),
                    app.notes,
                ])

        return str(output)

    def export_to_json(self, output_path: str, indent: int = 2) -> str:
        """
        Export data to JSON file.

        Args:
            output_path: Path for JSON file
            indent: JSON indentation level

        Returns:
            Path to created file
        """
        output = Path(output_path).expanduser()

        data = {
            'exported_at': datetime.now().isoformat(),
            'total_applications': len(self.get_all_applications()),
            'applications': [app.to_dict() for app in self.get_all_applications()],
        }

        with open(output, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=indent)

        return str(output)

    def update_notes(self, company: str, notes: str, append: bool = False) -> bool:
        """
        Update notes for a company.

        Args:
            company: Company name
            notes: New notes text
            append: If True, append to existing notes

        Returns:
            True if successful
        """
        row_index = self.find_company(company)
        if not row_index:
            return False

        notes_cell = self.worksheet.cell(row=row_index, column=COLUMNS['notes'])

        if append and notes_cell.value:
            notes_cell.value = f"{notes_cell.value}; {notes}"
        else:
            notes_cell.value = notes

        self._modified = True
        return True

    def clear_conflict(self, company: str) -> bool:
        """
        Clear conflict flag from a company's notes.

        Args:
            company: Company name

        Returns:
            True if successful
        """
        row_index = self.find_company(company)
        if not row_index:
            return False

        notes_cell = self.worksheet.cell(row=row_index, column=COLUMNS['notes'])
        current_notes = notes_cell.value or ""

        # Remove conflict notes
        import re
        cleaned = re.sub(r'Conflict:.*?(;|$)', '', current_notes)
        notes_cell.value = cleaned.strip('; ')

        # Remove conflict highlighting
        notes_cell.fill = PatternFill()  # Reset fill

        self._modified = True
        return True

    def manual_status_update(
        self,
        company: str,
        new_status: str,
        force: bool = False,
    ) -> ExcelUpdateResult:
        """
        Manually update status (bypasses hierarchy if force=True).

        Args:
            company: Company name
            new_status: New status to set
            force: If True, bypass hierarchy rules

        Returns:
            ExcelUpdateResult with operation details
        """
        row_index = self.find_company(company)
        if not row_index:
            return ExcelUpdateResult(
                success=False,
                message=f"Company not found: {company}",
            )

        current = self.get_application(row_index)

        if not force:
            # Check hierarchy
            update_result = can_update_status(current.status, new_status)
            if not update_result.allowed:
                return ExcelUpdateResult(
                    success=False,
                    is_conflict=True,
                    company=company,
                    old_status=current.status,
                    new_status=new_status,
                    message=f"Cannot update: {update_result.reason}",
                )

        # Update status
        self.worksheet.cell(row=row_index, column=COLUMNS['status'], value=new_status)
        self.worksheet.cell(row=row_index, column=COLUMNS['date_last'], value=datetime.now().strftime('%Y-%m-%d'))
        self._apply_status_formatting(row_index, new_status)

        self._modified = True

        return ExcelUpdateResult(
            success=True,
            is_update=True,
            row_index=row_index,
            company=company,
            old_status=current.status,
            new_status=new_status,
            message=f"Updated {company}: {current.status} -> {new_status}",
        )

    def close(self) -> None:
        """Close the workbook without saving."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None


# =============================================================================
# Helper Functions
# =============================================================================

def create_excel_storage(config: Dict[str, Any]) -> ExcelStorage:
    """
    Create ExcelStorage instance from configuration.

    Args:
        config: Configuration dictionary with excel settings

    Returns:
        Configured ExcelStorage instance
    """
    excel_config = config.get('excel', {})

    return ExcelStorage(
        file_path=excel_config.get('file_path', '~/job_applications.xlsx'),
        backup_dir=excel_config.get('backup_directory', '~/.emailagent/backups/'),
        auto_backup=excel_config.get('auto_backup', True),
        backup_retention_days=excel_config.get('backup_retention_days', 7),
    )


def format_summary_table(stats: Dict[str, Any]) -> str:
    """
    Format statistics as a summary table.

    Args:
        stats: Statistics dictionary from get_statistics()

    Returns:
        Formatted string table
    """
    lines = [
        "=" * 50,
        "JOB APPLICATION SUMMARY",
        "=" * 50,
        f"Total Companies: {stats['total_companies']}",
        "",
        "Status Breakdown:",
    ]

    for status, count in stats['status_counts'].items():
        lines.append(f"  {status}: {count}")

    lines.extend([
        "",
        "Confidence Levels:",
    ])

    for conf, count in stats['confidence_counts'].items():
        lines.append(f"  {conf}: {count}")

    if stats['conflict_count'] > 0:
        lines.extend([
            "",
            f"Conflicts: {stats['conflict_count']}",
        ])

    lines.append("=" * 50)

    return "\n".join(lines)
